from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from app.database import get_db
from app.models.sale import Sale, SaleItem
from app.models.product import Product
from app.deps import get_current_admin, get_current_cashier_or_admin

router = APIRouter(prefix="/export", tags=["export"])
NEGOCIO = "Heliud Refaccionaria"
PAYMENT = {"cash": "Efectivo", "card": "Tarjeta", "transfer": "Transferencia", "credit": "Crédito"}


# ── helpers ──────────────────────────────────────────────────────────────────

def get_sales_data(db, date_from, date_to):
    q = db.query(Sale)
    if date_from:
        q = q.filter(Sale.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.filter(Sale.created_at <= datetime.combine(date_to, datetime.max.time()))
    return q.order_by(Sale.created_at.desc()).all()


def sales_rows(sales):
    headers = ["#", "Fecha", "Cliente", "Total", "Descuento", "Pago", "Estado"]
    rows = []
    for s in sales:
        rows.append([
            s.id,
            s.created_at.strftime("%d/%m/%Y %H:%M"),
            s.customer.name if s.customer else "—",
            round(s.total, 2),
            round(s.discount, 2),
            PAYMENT.get(s.payment_method, s.payment_method),
            s.status,
        ])
    total_row = ["", "TOTAL", "", round(sum(s.total for s in sales), 2), "", "", ""]
    return headers, rows, total_row


def get_report_data(db, report):
    if report == "profit":
        results = (db.query(Product.name,
                            func.sum(SaleItem.quantity).label("qty"),
                            func.sum(SaleItem.subtotal).label("revenue"),
                            func.sum(SaleItem.quantity * Product.cost).label("cost"))
                   .join(SaleItem, SaleItem.product_id == Product.id)
                   .group_by(Product.id).all())
        headers = ["Producto", "Unidades", "Ingresos", "Costo", "Utilidad"]
        rows = [[r.name, r.qty, round(r.revenue, 2), round(r.cost or 0, 2),
                 round((r.revenue or 0) - (r.cost or 0), 2)] for r in results]
        total = ["TOTAL", sum(r[1] for r in rows), round(sum(r[2] for r in rows), 2),
                 round(sum(r[3] for r in rows), 2), round(sum(r[4] for r in rows), 2)]
        return headers, rows, total

    if report == "by_payment":
        results = (db.query(Sale.payment_method, func.count(Sale.id).label("count"),
                            func.sum(Sale.total).label("total"))
                   .group_by(Sale.payment_method).all())
        headers = ["Método de pago", "Ventas", "Total"]
        rows = [[PAYMENT.get(r.payment_method, r.payment_method), r.count, round(r.total, 2)]
                for r in results]
        total = ["TOTAL", sum(r[1] for r in rows), round(sum(r[2] for r in rows), 2)]
        return headers, rows, total

    # by_category
    results = (db.query(func.coalesce(func.cast(Product.category_id, db.bind.dialect.name == 'sqlite' and 'TEXT' or 'VARCHAR'), 'Sin categoría').label("cat"),
                        func.sum(SaleItem.quantity).label("qty"),
                        func.sum(SaleItem.subtotal).label("total"))
               .join(SaleItem, SaleItem.product_id == Product.id)
               .group_by(Product.category_id).all())
    headers = ["Categoría", "Unidades", "Total"]
    rows = [[r.cat or "Sin categoría", r.qty, round(r.total, 2)] for r in results]
    total = ["TOTAL", sum(r[1] for r in rows), round(sum(r[2] for r in rows), 2)]
    return headers, rows, total


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(title, headers, rows, total_row):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="E0E7FF")

    # Title row
    ws.append([f"{NEGOCIO} — {title}  ({date.today()})"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    # Headers
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row in rows:
        ws.append(row)

    # Total
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = total_font
        cell.fill = total_fill

    # Autofit columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF builder ───────────────────────────────────────────────────────────────

def build_pdf(title, headers, rows, total_row):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=40, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(NEGOCIO, styles["Title"]))
    elements.append(Paragraph(f"{title} — Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [headers] + [[str(v) for v in r] for r in rows] + [[str(v) for v in total_row]]
    col_count = len(headers)
    col_width = (A4[0] - 60) / col_count

    t = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F3F4F6")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E0E7FF")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf


# ── Endpoints ─────────────────────────────────────────────────────────────────

def excel_response(buf, filename):
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})

def pdf_response(buf, filename):
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/sales/excel")
def export_sales_excel(date_from: Optional[date] = None, date_to: Optional[date] = None,
                       db: Session = Depends(get_db), _=Depends(get_current_cashier_or_admin)):
    sales = get_sales_data(db, date_from, date_to)
    headers, rows, total = sales_rows(sales)
    return excel_response(build_excel("Historial de Ventas", headers, rows, total), f"ventas_{date.today()}.xlsx")

@router.get("/sales/pdf")
def export_sales_pdf(date_from: Optional[date] = None, date_to: Optional[date] = None,
                     db: Session = Depends(get_db), _=Depends(get_current_cashier_or_admin)):
    sales = get_sales_data(db, date_from, date_to)
    headers, rows, total = sales_rows(sales)
    return pdf_response(build_pdf("Historial de Ventas", headers, rows, total), f"ventas_{date.today()}.pdf")

@router.get("/reports/excel")
def export_report_excel(report: str = Query("profit", enum=["profit", "by_category", "by_payment"]),
                        db: Session = Depends(get_db), _=Depends(get_current_admin)):
    titles = {"profit": "Reporte de Utilidad", "by_category": "Ventas por Categoría", "by_payment": "Ventas por Método de Pago"}
    headers, rows, total = get_report_data(db, report)
    return excel_response(build_excel(titles[report], headers, rows, total), f"reporte_{report}_{date.today()}.xlsx")

@router.get("/reports/pdf")
def export_report_pdf(report: str = Query("profit", enum=["profit", "by_category", "by_payment"]),
                      db: Session = Depends(get_db), _=Depends(get_current_admin)):
    titles = {"profit": "Reporte de Utilidad", "by_category": "Ventas por Categoría", "by_payment": "Ventas por Método de Pago"}
    headers, rows, total = get_report_data(db, report)
    return pdf_response(build_pdf(titles[report], headers, rows, total), f"reporte_{report}_{date.today()}.pdf")
